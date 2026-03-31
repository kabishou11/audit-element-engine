from __future__ import annotations

from functools import lru_cache
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from app.models import AttachmentRule

RULE_HINT_KEYWORDS: dict[str, list[str]] = {
    "银行回单 / 电子回单": ["回单", "电子回单", "交易流水", "付款人账号", "收款人账号", "银行回单"],
    "发票 / 电子发票": ["发票", "发票号码", "开票日期", "价税合计", "税率", "发票专用章"],
    "合同 / 协议": ["合同", "协议", "甲方", "乙方", "签订", "合同期限"],
    "大额支出审批表": ["审批表", "申请单", "审批", "经办人", "两委意见", "村务监督委员会", "分管领导", "主要领导"],
    "拨付请示": ["请示", "拨付", "申请拨付", "资金拨付", "事由"],
    "党政和人大办公室办文单": ["办文单", "来文", "批示", "人大办公室", "办公室"],
    "报价单": ["报价单", "报价", "单价", "数量", "报价日期"],
    "经济组织内部结算凭证": ["结算凭证", "领款人", "证明人", "审核人", "内部结算"],
}


def _split_multi_value(raw_value: str) -> list[str]:
    parts = [part.strip() for part in raw_value.replace("；", ";").replace("，", ",").split(",")]
    if len(parts) == 1:
        parts = [part.strip() for part in raw_value.replace("；", ";").split(";")]
    return [part for part in parts if part]


def _split_aliases(attachment_type: str) -> list[str]:
    return [part.strip() for part in attachment_type.split("/") if part.strip()]


def _normalize_text(value: str) -> str:
    return value.replace(" ", "").replace("　", "").strip().lower()


def load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8")


@lru_cache
def load_rules(excel_path: Path) -> tuple[AttachmentRule, ...]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rules: list[AttachmentRule] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        attachment_type, attachment_description, audit_elements, audit_issue_references = row[:4]
        if not attachment_type:
            continue
        rule = AttachmentRule(
            attachment_type=str(attachment_type).strip(),
            attachment_description=str(attachment_description or "").strip(),
            ai_audit_elements=_split_multi_value(str(audit_elements or "")),
            audit_issue_references=_split_multi_value(str(audit_issue_references or "")),
            aliases=_split_aliases(str(attachment_type).strip()),
        )
        rules.append(rule)
    if not any(_normalize_text(rule.attachment_type) == _normalize_text("其他") for rule in rules):
        rules.append(
            AttachmentRule(
                attachment_type="其他",
                attachment_description="未在规则表标准附件类型中明确匹配的文档",
                ai_audit_elements=[],
                audit_issue_references=[],
                aliases=["其他"],
            )
        )
    return tuple(rules)


def build_rule_prompt_lines(rules: Iterable[AttachmentRule]) -> str:
    lines: list[str] = []
    for index, rule in enumerate(rules, start=1):
        lines.append(
            f"{index}. 附件类型: {rule.attachment_type}；描述: {rule.attachment_description}；"
            f"AI审计要素: {'、'.join(rule.ai_audit_elements) or '无'}；"
            f"审计对应问题参考: {'；'.join(rule.audit_issue_references) or '无'}"
        )
    return "\n".join(lines)


def _text_bigrams(text: str) -> set[str]:
    normalized = _normalize_text(text)
    return {normalized[index : index + 2] for index in range(len(normalized) - 1)} if len(normalized) >= 2 else set()


def _extract_phrases(text: str) -> list[str]:
    candidates = re.split(r"[、，,；;。:\s（）()]+", text)
    return [item.strip() for item in candidates if len(item.strip()) >= 2]


def build_rule_search_text(rule: AttachmentRule) -> str:
    return " ".join(
        [
            rule.attachment_type,
            rule.attachment_description,
            " ".join(rule.ai_audit_elements),
            " ".join(rule.aliases),
        ]
    )


def _alias_fragments(rule: AttachmentRule) -> set[str]:
    fragments: set[str] = set()
    for alias in [rule.attachment_type, *rule.aliases]:
        normalized = _normalize_text(alias)
        for index in range(len(normalized)):
            for width in range(2, 5):
                fragment = normalized[index : index + width]
                if len(fragment) >= 2:
                    fragments.add(fragment)
    return fragments


def score_rule(
    rule: AttachmentRule,
    source_text: str,
    alias_bonus: float,
    fragment_bonus: float,
    element_bonus: float,
    bigram_weight: float,
) -> tuple[float, dict[str, float]]:
    normalized_source = _normalize_text(source_text)
    search_text = build_rule_search_text(rule)
    source_bigrams = _text_bigrams(source_text)
    rule_bigrams = _text_bigrams(search_text)

    alias_hits = sum(1 for alias in rule.aliases if alias and _normalize_text(alias) in normalized_source)
    fragment_hits = sum(1 for fragment in _alias_fragments(rule) if fragment in normalized_source)
    element_hits = sum(1 for item in rule.ai_audit_elements if item and _normalize_text(item) in normalized_source)
    phrase_hits = sum(1 for item in _extract_phrases(rule.attachment_description) if _normalize_text(item) in normalized_source)
    hint_hits = sum(
        1
        for item in RULE_HINT_KEYWORDS.get(rule.attachment_type, [])
        if item and _normalize_text(item) in normalized_source
    )
    overlap_ratio = (len(source_bigrams & rule_bigrams) / max(len(rule_bigrams), 1)) if rule_bigrams else 0.0

    score = (
        alias_hits * alias_bonus
        + fragment_hits * fragment_bonus
        + element_hits * element_bonus
        + phrase_hits * 0.8
        + hint_hits * 1.6
        + overlap_ratio * bigram_weight
    )
    return score, {
        "alias_hits": float(alias_hits),
        "fragment_hits": float(fragment_hits),
        "element_hits": float(element_hits),
        "phrase_hits": float(phrase_hits),
        "hint_hits": float(hint_hits),
        "bigram_overlap": round(overlap_ratio, 4),
        "score": round(score, 4),
    }


def recall_candidate_rules(
    rules: Iterable[AttachmentRule],
    source_text: str,
    top_k: int,
    min_score: float,
    alias_bonus: float,
    fragment_bonus: float,
    element_bonus: float,
    bigram_weight: float,
) -> list[tuple[AttachmentRule, dict[str, float]]]:
    scored: list[tuple[AttachmentRule, dict[str, float]]] = []
    for rule in rules:
        score, details = score_rule(
            rule=rule,
            source_text=source_text,
            alias_bonus=alias_bonus,
            fragment_bonus=fragment_bonus,
            element_bonus=element_bonus,
            bigram_weight=bigram_weight,
        )
        if score >= min_score:
            scored.append((rule, details))
    scored.sort(key=lambda item: item[1]["score"], reverse=True)
    if scored:
        return scored[:top_k]
    fallback = []
    for rule in rules:
        fallback.append(
            (
                rule,
                {
                    "alias_hits": 0.0,
                    "fragment_hits": 0.0,
                    "element_hits": 0.0,
                    "phrase_hits": 0.0,
                    "hint_hits": 0.0,
                    "bigram_overlap": 0.0,
                    "score": 0.0,
                },
            )
        )
    return fallback[:top_k]


def match_rule_by_type(rules: Iterable[AttachmentRule], attachment_type: str) -> AttachmentRule:
    target = _normalize_text(attachment_type)
    for rule in rules:
        if _normalize_text(rule.attachment_type) == target:
            return rule
        if any(_normalize_text(alias) == target for alias in rule.aliases):
            return rule
    raise ValueError(f"无法在规则表中找到附件类型: {attachment_type}")
